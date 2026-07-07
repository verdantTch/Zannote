# -*- coding: utf-8 -*-
"""
Created on Wed Jul  1 11:14:34 2026
@author: hugoz
"""
import cv2
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def save_zannote_csv(image_path, points, output_folder):

    image = cv2.imread(str(image_path))
    height, width = image.shape[:2]
    image_name = Path(image_path).stem

    rows = []
    for i, (x, y, probability) in enumerate(points, start=1):
        rows.append({
            "image": image_name,
            "width": width,
            "height": height,
            "egg_id": i,
            "x": x,
            "y": y,
            "confidence": probability
        })

    df = pd.DataFrame(rows)
    output_file = Path(output_folder) / f"{image_name}.csv"
    df.to_csv(output_file, index=False)
    return output_file


def build_summary_rows(label_folder):
    """
    Reconstruit les lignes du résumé en relisant tous les CSV de
    label_folder, que ces CSV viennent d'une annotation manuelle ou
    d'une prédiction IA (même format de colonnes des deux côtés).

    Les images à 0 œuf restent affichées (egg_count=0) plutôt que d'être
    silencieusement absentes du tableau.
    """

    label_folder = Path(label_folder)
    rows = []

    for csv_file in sorted(label_folder.glob("*.csv")):

        df = pd.read_csv(csv_file)

        if df.empty:
            rows.append({
                "image": csv_file.stem,
                "egg_count": 0,
                "mean_probability": 0.0,
                "std_probability": 0.0
            })
            continue

        probabilities = (
            df["confidence"]
            if "confidence" in df.columns
            else pd.Series(np.ones(len(df)))
        )

        rows.append({
            "image": csv_file.stem,
            "egg_count": len(df),
            "mean_probability": round(float(probabilities.mean()), 3),
            "std_probability": round(float(probabilities.std(ddof=0)), 3)
        })

    return rows


def save_summary(rows, destination_folder):
    """
    Écrit et met en forme Summary.xlsx dans destination_folder à partir
    des lignes déjà calculées (mise en forme commune aux deux flux :
    en-tête coloré/gras, cellules centrées, colonnes ajustées).
    """

    destination_folder = Path(destination_folder)
    destination_folder.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(rows)
    output_file = destination_folder / "Summary.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        df.to_excel(writer, index=False)
        ws = writer.sheets["Sheet1"]

        fill = PatternFill(fill_type="solid", fgColor="11BED5")

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for column in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column
            )
            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = length + 4

        for column in ws.columns:
            for cell in column:
                cell.alignment = Alignment(horizontal="center")

    return output_file


def update_summary(label_folder, destination_folder):
    """
    Point d'entrée unique, utilisé à l'identique par CsvManager
    (annotation manuelle) et par Predictor (prédiction IA), pour garantir
    un Summary.xlsx strictement identique dans les deux cas.
    """

    rows = build_summary_rows(label_folder)
    return save_summary(rows, destination_folder)