# -*- coding: utf-8 -*-
"""
Created on Wed Jul  1 11:14:34 2026

@author: hugoz
"""

# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import cv2
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def save_zannote_csv(
    image_path,
    points,
    output_folder
):
    image = cv2.imread(str(image_path))
    height, width = image.shape[:2]

    image_name = Path(image_path).stem

    rows = []

    for i, (x, y, probability) in enumerate(points, start=1):

        rows.append({
            "image": image_name,
            "width": width,
            "height": height,
            "egg_id": i,
            "x": x,
            "y": y,
            "confidence": probability
        })

    df = pd.DataFrame(rows)

    output_file = Path(output_folder) / f"{image_name}.csv"

    df.to_csv(
        output_file,
        index=False
    )

    return output_file

    
def summary_row(
    image_name,
    points
):

    probabilities = [
        p[2]
        for p in points
    ]

    return {

        "image": image_name,

        "egg_count": len(points),

        "mean_probability":
            round(np.mean(probabilities), 3)
            if probabilities else 0,

        "std_probability":
            round(np.std(probabilities), 3)
            if probabilities else 0
    }
        
            
def save_summary(rows, output_folder):

    df = pd.DataFrame(rows)

    output_file = Path(output_folder) / "Summary.xlsx"

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False
        )

        ws = writer.sheets["Sheet1"]

        fill = PatternFill(
            fill_type="solid",
            fgColor="11BED5"
        )

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = fill

            cell.alignment = Alignment(
                horizontal="center"
            )

        for column in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column
            )

            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = length + 4

        for column in ws.columns:
            for cell in column:
                cell.alignment = Alignment(
                    horizontal="center"
                )

    return output_file